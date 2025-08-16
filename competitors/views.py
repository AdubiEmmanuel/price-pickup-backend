import csv
import io
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter, OrderingFilter
from .models import CompetitorPrice
from .serializers import CompetitorPriceSerializer

class CompetitorPriceViewSet(viewsets.ModelViewSet):
    queryset = CompetitorPrice.objects.all()
    serializer_class = CompetitorPriceSerializer
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['sku_category', 'sku_size', 'brand', 'is_unilever', 'location']
    search_fields = ['sku_name', 'brand']
    ordering_fields = ['created_at', 'sku_name', 'kd_case', 'kd_unit']

    @action(detail=False, methods=['get'])
    def get_category_choices(self, request):
        """
        Get available choices for SKU categories, size categories, and market types.
        """
        return Response({
            'sku_categories': dict(CompetitorPrice.SKU_CATEGORY_CHOICES),
            'sku_sizes': dict(CompetitorPrice.SKU_SIZE_CHOICES),
            'market_types': dict(CompetitorPrice.MARKET_CHOICES),
        })

    @action(detail=False, methods=['post'])
    def create_price_entry(self, request):
        """
        Create a new price entry with market-specific pricing.
        Expected payload:
        {
            "sku_category": "NUTRITION",
            "sku_size": "BULK PACK",
            "sku_name": "Product Name",
            "brand": "Brand Name",
            "market_type": "OPEN_MARKET",
            "price": 150.00,
            "is_unilever": true,
            "location": "Lagos"
        }
        """
        data = request.data.copy()

        # Validate required fields
        required_fields = ['sku_category', 'sku_size', 'sku_name', 'market_type', 'price']
        for field in required_fields:
            if not data.get(field):
                return Response(
                    {"error": f"{field.replace('_', ' ').title()} is required"},
                    status=status.HTTP_400_BAD_REQUEST
                )

        # Map market type to appropriate price field
        market_type = data.get('market_type')
        price = data.get('price')

        price_field_mapping = {
            'OPEN_MARKET': 'open_market_price',
            'NG': 'ng_price',
            'SMALL_SUPERMARKET': 'small_supermarket_price',
            'WHOLESALE': 'wholesale_price',
        }

        if market_type not in price_field_mapping:
            return Response(
                {"error": f"Invalid market type. Must be one of: {', '.join(price_field_mapping.keys())}"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Set the appropriate price field
        data[price_field_mapping[market_type]] = price

        # Remove the generic price and market_type fields as they're not model fields
        data.pop('price', None)
        data.pop('market_type', None)

        # Set source as FORM
        data['source'] = 'FORM'

        # Create the record
        serializer = self.get_serializer(data=data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post', 'patch'])
    def update_price(self, request, pk=None):
        """
        Update a single market-specific price for a SKU.
        Accepts: market_type (OPEN_MARKET|NG|SMALL_SUPERMARKET|WHOLESALE) and price (number).
        For records with source='CSV', this will clone and create a new FORM record as per serializer.update() logic.
        """
        market_type = request.data.get('market_type')
        price = request.data.get('price')
        if market_type is None or price is None:
            return Response({"error": "market_type and price are required"}, status=status.HTTP_400_BAD_REQUEST)

        mapping = {
            'OPEN_MARKET': 'open_market_price',
            'NG': 'ng_price',
            'SMALL_SUPERMARKET': 'small_supermarket_price',
            'WHOLESALE': 'wholesale_price',
        }
        if market_type not in mapping:
            return Response({"error": f"Invalid market type. Must be one of: {', '.join(mapping.keys())}"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            price_val = float(price)
        except (TypeError, ValueError):
            return Response({"error": "price must be a number"}, status=status.HTTP_400_BAD_REQUEST)

        partial_data = {mapping[market_type]: price_val}
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=partial_data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'])
    def upload(self, request):
        """
        Upload and process CSV/Excel files
        """
        if 'file' not in request.FILES:
            return Response({"error": "No file provided"}, status=status.HTTP_400_BAD_REQUEST)
        
        file = request.FILES['file']
        
        # Check file extension
        if not (file.name.endswith('.csv') or file.name.endswith('.xlsx')):
            return Response({"error": "File must be CSV or Excel format"}, status=status.HTTP_400_BAD_REQUEST)
        
        # Process CSV file
        if file.name.endswith('.csv'):
            try:
                decoded_file = file.read().decode('utf-8')
                io_string = io.StringIO(decoded_file)
                reader = csv.DictReader(io_string)
                
                # Map CSV headers to model fields
                field_mapping = {
                    'SKU Category': 'sku_category',
                    'SKU Size': 'sku_size',
                    'SKU Name/Brand': 'sku_name_brand',  # Split into sku_name and brand using known brands
                    'SKU Name': 'sku_name',              # Support separate SKU Name column
                    'Brand': 'brand',                    # Support separate Brand column
                    'KD Case': 'kd_case',
                    'KD Unit': 'kd_unit',
                    'KD Price/Gram': 'kd_price_gram',
                    'Wholesales price': 'wholesale_price',   # Some sheets use this header
                    'Wholesale price': 'wholesale_price',    # Accept common variant
                    'Wholesale Price': 'wholesale_price',
                    'Open Market Price': 'open_market_price',
                    'Open Market price': 'open_market_price',
                    'NG Price': 'ng_price',
                    'Small Supermarket Price': 'small_supermarket_price',
                }
                
                # Validate headers
                headers = reader.fieldnames
                if not headers or 'SKU Category' not in headers:
                    return Response({"error": "CSV format is invalid - SKU Category not found"}, 
                                   status=status.HTTP_400_BAD_REQUEST)
                
                # Process rows
                created_count = 0
                error_rows = []
                
                for row_num, row in enumerate(reader, start=2):  # Start at 2 to account for header row
                    try:
                        data = {}
                        for csv_field, model_field in field_mapping.items():
                            if csv_field in row:
                                value = row[csv_field].strip()
                                
                                # Handle the SKU Name/Brand field - split into name and brand
                                if model_field == 'sku_name_brand' and value:
                                    # Try to detect the brand from known BRAND_CHOICES within the string
                                    known_brands = [b for (b, _) in CompetitorPrice.BRAND_CHOICES]
                                    val_upper = value.upper()
                                    detected_brand = None
                                    for b in known_brands:
                                        if b in val_upper:
                                            detected_brand = b
                                            break
                                    if detected_brand:
                                        data['brand'] = detected_brand
                                        # Remove the detected brand token from the name (case-insensitive)
                                        data['sku_name'] = ' '.join([w for w in value.split() if w.upper() != detected_brand])
                                        data['sku_name'] = data['sku_name'].strip() or value
                                    else:
                                        # If we cannot detect brand, keep whole string as name and leave brand empty
                                        data['sku_name'] = value
                                        data['brand'] = None
                                # Handle numeric fields
                                elif model_field in ('kd_case', 'kd_unit', 'kd_price_gram', 
                                                    'wholesale_price', 'open_market_price', 
                                                    'ng_price', 'small_supermarket_price'):
                                    if value:
                                        try:
                                            data[model_field] = float(value.replace(',', ''))
                                        except ValueError:
                                            # Skip invalid numeric values
                                            pass
                                else:
                                    data[model_field] = value
                        
                        # Normalize category/size/brand against allowed choices
                        if data.get('sku_category'):
                            cat = str(data['sku_category']).strip().upper()
                            cat_syn = {'ORALS': 'ORAL CARE', 'DEOS': 'DEODORANT', 'DEODORANTS': 'DEODORANT'}
                            data['sku_category'] = cat_syn.get(cat, cat)
                        if data.get('sku_size'):
                            data['sku_size'] = str(data['sku_size']).strip().upper()
                        if data.get('brand'):
                            brand_map = {'CLOSEUP': 'CLOSE UP'}
                            b = str(data['brand']).strip().upper()
                            data['brand'] = brand_map.get(b, b)
                        # Set source to CSV (imports are treated as CSV source)
                        data['source'] = 'CSV'
                        # Heuristic for Unilever product can be adjusted; default to False
                        data['is_unilever'] = bool(data.get('is_unilever', False))
                        
                        # Create the record
                        serializer = self.get_serializer(data=data)
                        if serializer.is_valid():
                            serializer.save()
                            created_count += 1
                        else:
                            error_rows.append({
                                'row': row_num,
                                'errors': serializer.errors
                            })
                    except Exception as e:
                        error_rows.append({
                            'row': row_num,
                            'errors': str(e)
                        })
                
                return Response({
                    'message': f'Successfully created {created_count} records',
                    'errors': error_rows if error_rows else None
                }, status=status.HTTP_201_CREATED if created_count > 0 else status.HTTP_400_BAD_REQUEST)
                
            except Exception as e:
                return Response({"error": f"Error processing CSV: {str(e)}"}, 
                               status=status.HTTP_400_BAD_REQUEST)
        
        # For Excel files, process using openpyxl
        if file.name.endswith('.xlsx'):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(file)
                ws = wb.active
                
                # Map Excel headers to model fields
                field_mapping = {
                    'SKU Category': 'sku_category',
                    'SKU Size': 'sku_size',
                    'SKU Name': 'sku_name',
                    'Brand': 'brand',
                    'KD Case': 'kd_case',
                    'KD Unit': 'kd_unit',
                    'KD Price/Gram': 'kd_price_gram',
                    'Wholesales price': 'wholesale_price',
                    'Open Market Price': 'open_market_price',
                    'NG Price': 'ng_price',
                    'Small Supermarket Price': 'small_supermarket_price',
                }
                
                # Read headers from the first row
                headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                header_indices = {header: idx for idx, header in enumerate(headers)}
                
                # Validate headers
                if 'SKU Category' not in header_indices:
                    return Response({"error": "Excel format is invalid - SKU Category not found"}, status=status.HTTP_400_BAD_REQUEST)
                
                created_count = 0
                error_rows = []
                for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        data = {}
                        for excel_field, model_field in field_mapping.items():
                            idx = header_indices.get(excel_field)
                            if idx is not None:
                                value = row[idx]
                                if value is not None:
                                    value = str(value).strip()
                                else:
                                    value = ''
                                # Handle numeric fields
                                if model_field in ('kd_case', 'kd_unit', 'kd_price_gram', 'wholesale_price', 'open_market_price', 'ng_price', 'small_supermarket_price'):
                                    if value:
                                        try:
                                            data[model_field] = float(value.replace(',', ''))
                                        except Exception:
                                            pass
                                else:
                                    data[model_field] = value
                        # Normalize category/size/brand
                        if data.get('sku_category'):
                            cat = str(data['sku_category']).strip().upper()
                            cat_syn = {'ORALS': 'ORAL CARE', 'DEOS': 'DEODORANT', 'DEODORANTS': 'DEODORANT'}
                            data['sku_category'] = cat_syn.get(cat, cat)
                        if data.get('sku_size'):
                            data['sku_size'] = str(data['sku_size']).strip().upper()
                        if data.get('brand'):
                            brand_map = {'CLOSEUP': 'CLOSE UP'}
                            b = str(data['brand']).strip().upper()
                            data['brand'] = brand_map.get(b, b)
                        # Treat EXCEL imports as CSV source to satisfy model choices
                        data['source'] = 'CSV'
                        data['is_unilever'] = bool(data.get('is_unilever', False))
                        serializer = self.get_serializer(data=data)
                        if serializer.is_valid():
                            serializer.save()
                            created_count += 1
                        else:
                            error_rows.append({'row': row_num, 'errors': serializer.errors})
                    except Exception as e:
                        error_rows.append({'row': row_num, 'errors': str(e)})
                return Response({
                    'message': f'Successfully created {created_count} records',
                    'errors': error_rows if error_rows else None
                }, status=status.HTTP_201_CREATED if created_count > 0 else status.HTTP_400_BAD_REQUEST)
            except Exception as e:
                return Response({"error": f"Error processing Excel: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get'])
    def distinct_skus(self, request):
        """
        Return a distinct list of SKU entries for populating dashboards/forms.
        Applies the same filters as list endpoint.
        Response items: sku_name, brand, sku_category, sku_size, latest_id
        """
        filtered = self.filter_queryset(self.get_queryset())
        # Use values + distinct to get unique combinations
        rows = (filtered
                .values('sku_name', 'brand', 'sku_category', 'sku_size')
                .distinct())
        # Optionally fetch a latest id for each combo to update later
        results = []
        for row in rows:
            ref = (filtered.filter(
                sku_name=row['sku_name'],
                brand=row['brand'],
                sku_category=row['sku_category'],
                sku_size=row['sku_size']
            ).order_by('-created_at').first())
            row['latest_id'] = ref.id if ref else None
            results.append(row)
        return Response(results, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'])
    def search_skus(self, request):
        """
        Lightweight SKU search for typeahead.
        Query param: q (matches sku_name or brand, case-insensitive)
        Returns up to 50 distinct results.
        """
        q = request.query_params.get('q', '').strip()
        qs = self.get_queryset()
        if q:
            from django.db.models import Q
            qs = qs.filter(Q(sku_name__icontains=q) | Q(brand__icontains=q))
        rows = (qs.values('sku_name', 'brand', 'sku_category', 'sku_size')
                  .distinct()[:50])
        return Response(list(rows), status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'])
    def export(self, request):
        """
        Export data in CSV format
        """
        format_type = request.query_params.get('format', 'csv')
        
        if format_type.lower() != 'csv':
            return Response({"error": f"Format {format_type} not supported"}, 
                           status=status.HTTP_400_BAD_REQUEST)
        
        # Apply filters if provided
        queryset = self.filter_queryset(self.get_queryset())
        
        # Create CSV response
        response = Response(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="competitor_prices.csv"'
        
        # Write CSV data
        writer = csv.writer(response)
        
        # Write header row
        writer.writerow([
            'SKU Category', 
            'SKU Size', 
            'SKU Name/Brand',
            'KD Case',
            'KD Unit',
            'KD Price/Gram',
            'Wholesales price',
            'Open Market Price',
            'NG Price',
            'Small Supermarket Price',
            'Is Unilever',
            'Location',
            'Created At'
        ])
        
        # Write data rows
        for item in queryset:
            writer.writerow([
                item.sku_category,
                item.sku_size,
                f"{item.sku_name} {item.brand}".strip(),
                item.kd_case,
                item.kd_unit,
                item.kd_price_gram,
                item.wholesale_price,
                item.open_market_price,
                item.ng_price,
                item.small_supermarket_price,
                'Yes' if item.is_unilever else 'No',
                item.location,
                item.created_at.strftime('%Y-%m-%d %H:%M:%S')
            ])
        
        return response



